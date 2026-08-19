"""Synthetic robustness tests for Structure Analyzer, without Mongo access."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
import json
import os
from time import perf_counter

import pytest
from openpyxl import Workbook

from structure_analyzer import analyze_structure

pytestmark = pytest.mark.unit


def _require_test_environment() -> None:
    assert os.environ.get("TEST_DB_NAME") == "proposta_ja_test"
    assert os.environ.get("TEST_MONGO_URL")


def _write_xlsx_and_records(tmp_path: Path, filename: str, sheets: list[dict]) -> tuple[list[dict], list[dict]]:
    workbook = Workbook()
    workbook.remove(workbook.active)
    entries = []
    records = []
    for spec in sheets:
        sheet = workbook.create_sheet(spec["name"])
        headers = spec.get("headers", [])
        header_row = spec.get("header_row", 1)
        title_rows = spec.get("title_rows", [])
        for row_index, values in enumerate(title_rows, start=1):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
        if headers:
            for column_index, value in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=column_index, value=value)
            for offset, values in enumerate(spec.get("rows", []), start=1):
                row_number = header_row + offset
                for column_index, value in enumerate(values, start=1):
                    sheet.cell(row=row_number, column=column_index, value=value)
                record = {
                    "source_sheet": spec["name"],
                    "source_row": row_number,
                    "original_record_json": {
                        header: (values[index] if index < len(values) else "")
                        for index, header in enumerate(headers)
                    },
                    "raw_metadata": {
                        "headers": headers[:],
                        "header_row": header_row,
                        "header_detection_method": "structural",
                        "header_detection_status": "DETECTED",
                        "sheet_rows": header_row + len(spec.get("rows", [])),
                        "sheet_columns": len(headers),
                    },
                }
                records.append(record)
        entries.append({
            "sheet": spec["name"],
            "rows": 0 if not headers else header_row + len(spec.get("rows", [])),
            "columns": len(headers),
            "headers": headers[:],
            "header_row": header_row if headers else None,
            "header_detection_method": "structural",
            "header_detection_status": "DETECTED" if headers else "NOT_APPLICABLE",
        })
    workbook.save(tmp_path / filename)
    return entries, records


def _analyze(tmp_path: Path, filename: str, sheets: list[dict]) -> tuple[dict, float, list[dict]]:
    _require_test_environment()
    entries, records = _write_xlsx_and_records(tmp_path, filename, sheets)
    before = deepcopy(records)
    started = perf_counter()
    profile = analyze_structure(
        {
            "filename": filename,
            "file_type": "XLSX",
            "file_size": (tmp_path / filename).stat().st_size,
            "checksum": "synthetic-checksum",
        },
        records,
        entries,
    )
    elapsed = perf_counter() - started
    assert records == before
    return profile, elapsed, records


def _sheet(profile: dict, name: str) -> dict:
    return next(sheet for sheet in profile["sheets"] if sheet["sheet_name"] == name)


def _warning_codes(profile: dict) -> set[str]:
    return {warning["code"] for warning in profile["warnings"]}


def _report(label: str, profile: dict, elapsed: float) -> None:
    stats = profile["global_statistics"]
    print(
        f"{label}: PASS | {elapsed:.4f}s | sheets={stats['total_sheets']} "
        f"columns={stats['total_columns']} rows={stats['total_rows']} "
        f"warnings={sorted(_warning_codes(profile))}"
    )


def test_stress_a_simple_budget(tmp_path):
    rows = [
        [f"Cliente {index}", f"00.000.000/0001-{index:02d}", f"SKU-{index:04d}", "Material", index, "R$ 10,50", f"R$ {index * 10},00"]
        for index in range(1, 21)
    ]
    profile, elapsed, _ = _analyze(tmp_path, "a.xlsx", [{"name": "Orçamentos", "headers": ["Cliente", "CNPJ", "Código", "Descrição", "Qtd", "Vlr Unit.", "Total"], "rows": rows}])
    sheet = _sheet(profile, "Orçamentos")
    assert sheet["structure_status"] == "TABULAR"
    assert sheet["data_row_count"] == 20
    types = {column["source_name"]: column["data_type"] for column in sheet["columns"]}
    assert types["Qtd"] == "INTEGER"
    assert types["Vlr Unit."] == "CURRENCY_LIKE"
    assert "CNPJ_LIKE" in next(column for column in sheet["columns"] if column["source_name"] == "CNPJ")["pattern_flags"]
    _report("A", profile, elapsed)


def test_stress_b_nonstandard_names_do_not_map_semantics(tmp_path):
    rows = [[f"Empresa {index}", f"12.345.678/0001-{index:02d}", f"SKU-{index:04d}", "Produto", index, 10.25, 20.50] for index in range(1, 6)]
    profile, elapsed, _ = _analyze(tmp_path, "b.xlsx", [{"name": "Pedidos", "headers": ["Razão Social", "Documento", "SKU", "Produto", "Quantidade", "Preço", "Valor Final"], "rows": rows}])
    sheet = _sheet(profile, "Pedidos")
    columns = {column["source_name"]: column for column in sheet["columns"]}
    assert columns["Razão Social"]["semantic_meaning"] == "unknown"
    assert columns["SKU"]["potential_identifier"] is True
    assert columns["Quantidade"]["data_type"] == "INTEGER"
    assert columns["Preço"]["data_type"] == "DECIMAL"
    assert "CNPJ_LIKE" in columns["Documento"]["pattern_flags"]
    _report("B", profile, elapsed)


def test_stress_c_title_rows_and_header_line_five(tmp_path):
    rows = [["ACME", "ORC-001", "MAT-001", "Aço", "UN", 3, "100,00", "300,00"]]
    profile, elapsed, _ = _analyze(tmp_path, "c.xlsx", [{
        "name": "Relatório",
        "header_row": 5,
        "title_rows": [["RELATÓRIO COMERCIAL"], ["EMPRESA TESTE"], ["Emitido em 14/08/2026"], []],
        "headers": ["CLIENTE", "ORÇAMENTO Nº", "ITEM", "DESCRIÇÃO MATERIAL", "UN.", "QTD.", "PREÇO UNITÁRIO", "TOTAL"],
        "rows": rows,
    }])
    sheet = _sheet(profile, "Relatório")
    assert sheet["header_row"] == 5
    assert sheet["header_detection_status"] == "CONFIDENT"
    assert sheet["data_row_count"] == 1
    _report("C", profile, elapsed)


def test_stress_d_multiple_commercial_sheets(tmp_path):
    specs = [
        {"name": "Clientes", "headers": ["CNPJ", "Razão Social"], "rows": [["00.000.000/0001-00", "ACME"]]},
        {"name": "Produtos", "headers": ["Código", "Descrição"], "rows": [["SKU-001", "Aço"]]},
        {"name": "Orçamentos", "headers": ["CNPJ", "Orçamento", "Valor"], "rows": [["00.000.000/0001-00", "ORC-001", "100,00"]]},
        {"name": "Financeiro", "headers": ["Data", "Percentual", "Saldo"], "rows": [["2026-08-14", "10%", "1000.00"]]},
    ]
    profile, elapsed, _ = _analyze(tmp_path, "d.xlsx", specs)
    assert profile["global_statistics"]["total_sheets"] == 4
    assert profile["global_statistics"]["tabular_sheets"] == 4
    assert any(hint["hint_type"] == "potential_shared_key" for hint in profile["sheet_relationship_hints"])
    _report("D", profile, elapsed)


def test_stress_e_imperfect_data_warnings(tmp_path):
    rows = []
    for index in range(12):
        rows.append(["" if index < 10 else "ACME", "R$ 10,00" if index % 2 else "invalid", "10" if index % 3 else "unknown", "2026-08-14" if index % 2 else "14/08/2026", "DUP-001"])
    profile, elapsed, _ = _analyze(tmp_path, "e.xlsx", [{"name": "Imperfeito", "headers": ["Parcial", "Preço", "Quantidade", "Data", "Código"], "rows": rows}])
    codes = _warning_codes(profile)
    assert "MIXED_DATA_TYPES" in codes
    assert "HIGH_EMPTY_RATIO" in codes
    assert "HIGH_DUPLICATION" in codes
    _report("E", profile, elapsed)


def test_stress_f_duplicate_columns_are_reported(tmp_path):
    profile, elapsed, _ = _analyze(tmp_path, "f.xlsx", [{
        "name": "Duplicadas",
        "headers": ["Código", "Descrição", "Código", "Valor"],
        "rows": [["A-001", "Material", "A-001", "10,00"]],
    }])
    sheet = _sheet(profile, "Duplicadas")
    assert sheet["column_count"] == 4
    assert "DUPLICATE_COLUMN_NAMES" in _warning_codes(profile)
    _report("F", profile, elapsed)


def test_stress_g_empty_sheet_has_no_records(tmp_path):
    specs = [
        {"name": "Clientes", "headers": ["CNPJ"], "rows": [["00.000.000/0001-00"]]},
        {"name": "Produtos", "headers": ["Código"], "rows": [["SKU-001"]]},
        {"name": "AbaVazia", "headers": [], "rows": []},
    ]
    profile, elapsed, records = _analyze(tmp_path, "g.xlsx", specs)
    empty = _sheet(profile, "AbaVazia")
    assert empty["structure_status"] == "EMPTY"
    assert not any(record["source_sheet"] == "AbaVazia" for record in records)
    _report("G", profile, elapsed)


def test_stress_h_shared_key_hints_only(tmp_path):
    specs = [
        {"name": "Clientes", "headers": ["CNPJ", "Razão Social"], "rows": [["00.000.000/0001-00", "ACME"]]},
        {"name": "Orçamentos", "headers": ["CNPJ", "Orçamento", "Valor"], "rows": [["00.000.000/0001-00", "ORC-1", "10,00"]]},
        {"name": "Produtos", "headers": ["Código", "Descrição"], "rows": [["SKU-1", "Aço"]]},
        {"name": "Itens", "headers": ["Código", "Descrição", "Quantidade"], "rows": [["SKU-1", "Aço", 2]]},
    ]
    profile, elapsed, _ = _analyze(tmp_path, "h.xlsx", specs)
    hints = profile["sheet_relationship_hints"]
    assert hints
    assert all(hint["hint_type"] == "potential_shared_key" for hint in hints)
    assert "client_id" not in json.dumps(profile)
    _report("H", profile, elapsed)


def test_stress_i_unfamiliar_headers_still_describe_structure(tmp_path):
    rows = [[f"MAT-{index:03d}", "Aço", "UN", index, 10.0, index * 10.0] for index in range(1, 5)]
    profile, elapsed, _ = _analyze(tmp_path, "i.xlsx", [{"name": "Materiais", "headers": ["COD.", "DESC. MAT.", "UNID.", "QTDE.", "UNIT.", "TOTAL"], "rows": rows}])
    sheet = _sheet(profile, "Materiais")
    assert sheet["structure_status"] == "TABULAR"
    assert sheet["column_count"] == 6
    assert all(column["semantic_meaning"] == "unknown" for column in sheet["columns"])
    _report("I", profile, elapsed)


def test_stress_j_ten_thousand_rows_twenty_columns(tmp_path):
    headers = [f"Coluna {index:02d}" for index in range(20)]
    rows = [[f"SKU-{row:05d}" if column == 0 else (row if column % 3 == 0 else f"valor-{row}-{column}") for column in range(20)] for row in range(10000)]
    profile, elapsed, records = _analyze(tmp_path, "j.xlsx", [{"name": "Grande", "headers": headers, "rows": rows}])
    sheet = _sheet(profile, "Grande")
    serialized_profile = json.dumps(profile, ensure_ascii=False)
    assert sheet["data_row_count"] == 10000
    assert sheet["column_count"] == 20
    assert all(len(column["sample_values"]) <= 5 for column in sheet["columns"])
    assert len(serialized_profile) < 100000
    assert len(records) == 10000
    _report("J", profile, elapsed)
    print(f"J_PROFILE_BYTES: {len(serialized_profile)}")
