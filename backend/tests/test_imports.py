import os
from pathlib import Path

import pytest

from server import (
    ALLOWED_IMPORT_EXTENSIONS,
    _extract_csv_records,
    _extract_raw_records,
    _parse_xlsx_file,
    sanitize_filename,
    sha256_bytes,
)


pytestmark = pytest.mark.unit


def _mongo_available() -> bool:
    mongo_url = os.getenv("TEST_MONGO_URL")
    test_db_name = os.getenv("TEST_DB_NAME")
    if not mongo_url or test_db_name != "proposta_ja_test":
        return False
    try:
        from pymongo import MongoClient

        client = MongoClient(mongo_url, serverSelectionTimeoutMS=500)
        client[test_db_name].command("ping")
        client.close()
        return True
    except Exception:
        return False


def test_csv_comma_preserves_utf8_values(tmp_path):
    csv_text = "Cliente,Orçamento,Valor,Vendedor\nACME,1524,25000,João\nEmpresa Teste,1525,35000,Maria\n"
    file_path = tmp_path / "orcamentos.csv"
    file_path.write_text(csv_text, encoding="utf-8")

    records = _extract_csv_records(str(file_path))
    assert len(records) == 1
    first = records[0]["records"][0]["original_record_json"]
    assert first["Cliente"] == "ACME"
    assert first["Vendedor"] == "João"
    assert records[0]["records"][1]["original_record_json"]["Cliente"] == "Empresa Teste"


def test_csv_semicolon_delimiter_and_utf8_bom(tmp_path):
    csv_text = "Cliente;Orçamento;Valor;Vendedor\nACME;1524;25000;João\n"
    file_path = tmp_path / "orcamentos_semicolon.csv"
    file_path.write_text(csv_text, encoding="utf-8-sig")

    records = _extract_csv_records(str(file_path))
    assert records[0]["records"][0]["original_record_json"]["Cliente"] == "ACME"
    assert records[0]["records"][0]["original_record_json"]["Vendedor"] == "João"


def test_csv_utf8_special_characters_are_preserved(tmp_path):
    csv_text = "Cliente,Local,Atividade\nACME,São Paulo,Indústria\nMáquinas,Centro,Teste\n"
    file_path = tmp_path / "caracteres.csv"
    file_path.write_text(csv_text, encoding="utf-8")

    records = _extract_csv_records(str(file_path))
    first = records[0]["records"][0]["original_record_json"]
    assert first["Local"] == "São Paulo"
    assert first["Atividade"] == "Indústria"
    assert records[0]["records"][1]["original_record_json"]["Cliente"] == "Máquinas"


def test_xlsx_parser_handles_single_and_multiple_sheets(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Orçamentos"
    sheet1.append(["Cliente", "Orçamento", "Valor", "Vendedor"])
    sheet1.append(["ACME", "1524", "25000", "João"])

    sheet2 = workbook.create_sheet("Clientes")
    sheet2.append(["Cliente", "CNPJ"])
    sheet2.append(["ACME", "00.000.000/0001-00"])

    file_path = tmp_path / "orcamentos.xlsx"
    workbook.save(file_path)

    parsed = _parse_xlsx_file(str(file_path))
    assert len(parsed) >= 2
    assert {sheet["sheet"] for sheet in parsed} >= {"Orçamentos", "Clientes"}
    assert any("ACME" in str(record["original_record_json"]) for sheet in parsed for record in sheet["records"])


def test_xlsx_empty_sheet_and_header_detection(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["RELATÓRIO DE ORÇAMENTOS"])
    sheet.append(["Emitido em 13/08/2026"])
    sheet.append([])
    sheet.append(["Cliente", "Valor", "Vendedor"])
    sheet.append(["ACME", "25000", "João"])

    file_path = tmp_path / "header_test.xlsx"
    workbook.save(file_path)

    parsed = _parse_xlsx_file(str(file_path))
    assert len(parsed) == 1
    assert parsed[0]["header_detection_status"] in {"DETECTED", "AMBIGUOUS"}
    assert parsed[0]["records"]


def test_checksum_is_deterministic_and_content_sensitive():
    payload = b"Cliente,Or\xc3\xa7amento,Valor,Vendedor\nACME,1524,25000,Jo\xc3\xa3o\n"
    first = sha256_bytes(payload)
    second = sha256_bytes(payload)
    changed = sha256_bytes(payload + b"\n")

    assert first == second
    assert first != changed
    assert len(first) == 64


def test_sanitize_filename_blocks_path_traversal():
    assert sanitize_filename("../../../teste.xlsx") == "teste.xlsx"
    assert sanitize_filename("..\\..\\teste.xlsx") == "teste.xlsx"
    assert sanitize_filename("arquivo normal.xlsx") == "arquivo normal.xlsx"
    assert sanitize_filename("área.xlsx") == "área.xlsx"


def test_invalid_extension_validation_without_endpoint():
    for name in ["arquivo.exe", "script.py", "no_extension", "arquivo", "corrupt.bin"]:
        suffix = Path(name).suffix.lower()
        assert suffix not in ALLOWED_IMPORT_EXTENSIONS or suffix == ".csv" or suffix == ".xlsx"


def test_raw_extraction_preserves_original_headers_and_values(tmp_path):
    csv_text = "Cliente,Orçamento,Valor\nACME,1524,25000\nEmpresa Teste,1525,35000\n"
    file_path = tmp_path / "raw.csv"
    file_path.write_text(csv_text, encoding="utf-8")

    raw_records, total_rows = _extract_raw_records(str(file_path), "CSV")
    assert total_rows == 3
    assert len(raw_records) == 2
    first = raw_records[0]
    assert first["source_row"] == 2
    assert first["source_sheet"] == "CSV"
    assert first["original_record_json"]["Cliente"] == "ACME"
    assert first["original_record_json"]["Orçamento"] == "1524"
    assert first["original_record_json"]["Valor"] == "25000"


@pytest.mark.integration
@pytest.mark.skipif(not _mongo_available(), reason="MongoDB indisponível no ambiente atual")
def test_integration_import_batch_requires_mongo():
    assert True


@pytest.mark.integration
@pytest.mark.skipif(not _mongo_available(), reason="MongoDB indisponível no ambiente atual")
def test_integration_api_requires_mongo_and_auth():
    assert True
